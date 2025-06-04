import streamlit as st
from src import config as con
from src import load_data as load
from src import select_input as input
from openpyxl.styles import PatternFill


# --- COLOUR SCHEME ---
st.markdown("""
    <style>
    div.stButton > button {
        background-color: #f0f0f0 !important;
        color: black !important;
        border: 1px solid #ccc !important;
        padding: 0.5em 1.2em !important;
        border-radius: 6px !important;
        font-size: 1rem !important;
    }

    div.stButton > button:hover {
        background-color: #d6e4ff !important;  /* light blue */
        color: black !important;
        border-color: #a0c4ff !important;
    }
    </style>
""", unsafe_allow_html=True)


# --- MAIN APP ---
st.title("🇫🇷 French Verb Conjugation Trainer")

wb = load.load_workbook(con.EXCEL_FILE)
ws_input = wb["UserInput"]
ws_solution = wb["Solutions"]

if "current_task" not in st.session_state:
    row, col, verb, prompt = input.get_random_task(ws_solution)
    st.session_state.current_task = {
        "row": row,
        "col": col,
        "verb": verb,
        "prompt": prompt
    }
else:
    task = st.session_state.current_task
    row = task["row"]
    col = task["col"]
    verb = task["verb"]
    prompt = task["prompt"]

if st.session_state.get("clear_input"):
    st.session_state.user_input = ""
    st.session_state.clear_input = False

if row is None:
    st.success("🎉 All verbs have been completed!")
else:
    st.subheader(f"Verb: **{verb}**")
    st.write(f"Conjugate for: **{prompt}**")

    user_input = st.text_input("Your conjugation:", key="user_input")

    if st.button("Check answer"):
        # Fetch the correct answer from the solution sheet
        correct_answer = str(ws_solution[f"{col}{row}"].value).strip()
        user_input_clean = user_input.strip()

        # Save user input to the input sheet
        cell = ws_input[f"{col}{row}"] # TODO: Check correct cell referencing for UserInput sheet
        cell.value = user_input_clean

        # Define fill styles
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Compare and apply style
        if user_input_clean.lower() == correct_answer.lower():
            # TODO: Check on past participle correct forms (", ...")
            st.success("✅ Correct!")
            cell.fill = green_fill
        else:
            st.error(f"❌ Incorrect. Correct answer: **{correct_answer}**")
            cell.fill = red_fill

        # Check if full row is complete
        filled = all(ws_input[f"{c}{row}"].value not in [None, ""] for c in con.CONJUGATION_COLS)
        if filled:
            ws_input[f"{con.STATUS_COL}{row}"].value = "True"

        wb.save(con.EXCEL_FILE)

        # ✅ Clear the input field AFTER saving and feedback
        st.session_state.clear_input = True

if st.button("Next verb"):
    st.session_state.pop("current_task", None)
    st.rerun()


# TODO: Add that the correct solution is hidden if the answer was wrong, so you can retry (that needs logging of your wrong answers in a separate script then, that we could display on a second tab)
# TODO: Allow accent's to be omitted for the word to be correct?
# TODO: Set possible filters upfront (e.g. only -er/ir/-... verbs, specific tenses, ...)
# TODO: Display tenses and pronouns not in text form but on some kind of continuum /visual of all possibilities
# TODO: Not use the English translation as an input, but use it for a hidden field that shows the translation so you can also practice your vocabulary -> potential for a third tab with vocac trainer
