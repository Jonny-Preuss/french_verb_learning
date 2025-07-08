from openpyxl.styles import PatternFill
import streamlit as st

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def check_user_input(user_input, correct_answer, cell):
    user_input_clean = user_input.strip()
    correct_answer_clean = correct_answer.strip()
    
    is_correct = user_input_clean.lower() == correct_answer_clean.lower()
    
    cell.value = user_input_clean
    cell.fill = green_fill if is_correct else red_fill
    
    return is_correct, user_input_clean