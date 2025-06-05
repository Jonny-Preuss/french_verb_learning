import pandas as pd
from openpyxl import load_workbook
import time
import streamlit as st
from src import config as con


# --- FUNCTIONS ---
@st.cache_data
def load_dataframe():
    df = pd.read_excel(con.EXCEL_FILE, header=[0, 1], engine="openpyxl")
    return df

def load_workbook_sheet():
    wb = load_workbook(con.EXCEL_FILE)
    ws = wb.active
    return wb, ws


def safe_load_workbook(path, retries=3, delay=0.5):
    for i in range(retries):
        try:
            return load_workbook(path)
        except EOFError:
            time.sleep(delay)
    st.error("⚠️ Could not read the Excel file. Please make sure it's not open elsewhere.")
    return None