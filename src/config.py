from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
EXCEL_FILE = "data/Top_1000_verbs_French_USE_test.xlsx"
START_ROW = 3
VERB_COL = "B"
TRANSLATION_COL = "C"
FILTER_COL = "F"
STATUS_COL = "AW"
CONJUGATION_COLS = [get_column_letter(i) for i in range(6, 48)]  # Columns G to AV


# ---- COLOUR SCHEME ------
base = "light"
primaryColor = "#1f77b4"  # Streamlit's default blue


# ---- TENSE MAPPING ----
PRESENT_COLS = [get_column_letter(i) for i in range(15, 21)]  # Columns O to T
IMPARFAIT_COLS = [get_column_letter(i) for i in range(21, 27)]  # Columns U to Z
FUTUR_COLS = [get_column_letter(i) for i in range(28, 34)]  # Columns AB to AG
SUBJONCTIF_COLS = [get_column_letter(i) for i in range(35, 41)]  # Columns AI to AN
CONDITIONNEL_COLS = [get_column_letter(i) for i in range(42, 48)]  # Columns AP to AU
IMPERATIF_COLS = [get_column_letter(i) for i in range(12, 15)]  # Columns L to N
OTHER_COLS = ["G", "H", "I", "J", "K", "AA", "AH", "AO", "AV"]

TENSE_COL_MAP = {
    # "All": CONJUGATION_COLS,
    "Présent": PRESENT_COLS,
    "Imparfait": IMPARFAIT_COLS,
    "Futur": FUTUR_COLS,
    "Subjonctif": SUBJONCTIF_COLS,
    "Conditionnel": CONDITIONNEL_COLS,
    "Impératif": IMPERATIF_COLS,
    "Autres": OTHER_COLS,
}

TENSE_OPTIONS = ["(Random)"] + list(TENSE_COL_MAP.keys())