import streamlit as st
from src import config as con
from src import load_data as load
from src import select_input as input
from src.session import init_session_state 
from src.checking import check_user_input
from src.logging_attempts import log_incorrect_attempt
from openpyxl.styles import PatternFill
from datetime import datetime
import pandas as pd
import os


# --- COLOUR SCHEME --------
st.markdown("""
    <style>
    div.stButton > button {
        background-color: #f0f0f0 !important;
        color: black !important;
        border: 1px solid #ccc !important;
        padding: 0.5em 1.2em !important;
        border-radius: 6px !important;
        font-size: 1rem !important;
    }

    div.stButton > button:hover {
        background-color: #d6e4ff !important;  /* light blue */
        color: black !important;
        border-color: #a0c4ff !important;
    }
    </style>
""", unsafe_allow_html=True)

# --------- MAIN APP ---------
init_session_state()


# --------- APP TITLE ---------
st.title("🇫🇷 French Verb Conjugation Trainer")


# --------- LOAD WORKBOOK ---------
wb = load.safe_load_workbook(con.EXCEL_FILE)
if wb is None:
    st.stop()
ws_input = wb["UserInput"]
ws_solution = wb["Solutions"]


# --- FILTER UI SIDEBAR ---
filter_options = set()
for r in range(con.START_ROW, ws_solution.max_row + 1):
    val = ws_solution[f"{con.FILTER_COL}{r}"].value
    if val: filter_options.add(val.strip())

filter_options = sorted(filter_options)
with st.sidebar:
    selected_filter = st.radio("🔍 Filter by verb group:", ["(All)"] + filter_options, horizontal=False)

    st.markdown("---")

    if st.checkbox("Select all tenses", key="all_tenses"):
        print("ALL TENSES SELECTED")
        selected_tenses = list(con.TENSES.keys())
        
    selected_tenses = st.multiselect(
        "⏱️ Filter by tense:",
        options=list(con.TENSES.keys()),
        default=["Présent"]
    )
    print(selected_tenses)

    # Flatten the selected columns
    selected_cols = [col for tense in selected_tenses for col in con.TENSES[tense]]
    print(selected_cols)
    print(st.session_state)


# Reset task if filter changes
if "last_filter" not in st.session_state:
    st.session_state["last_filter"] = selected_filter

# if selected_filter != st.session_state["last_filter"]:
#     st.session_state["last_filter"] = selected_filter
#     st.session_state.pop("current_task", None)
#     st.session_state.reset_input = True
#     st.rerun()

if "last_filter" not in st.session_state:
    st.session_state["last_filter"] = selected_filter

# If the tense selection changed, reset current task
if "last_tense_selection" not in st.session_state:
    print("last_tense_selected not in st.session_state")
    st.session_state["last_tense_selection"] = selected_tenses
    print(selected_tenses)
# if selected_tenses != st.session_state["last_tense_selection"]:
#     print("selected_tenses not equal to session_state")
#     st.session_state["last_tense_selection"] = selected_tenses
#     st.session_state.pop("current_task", None)
#     st.session_state.reset_input = True
#     st.rerun()
#     print(selected_tenses)

if (
    selected_filter != st.session_state["last_filter"]
    or selected_tenses != st.session_state["last_tense_selection"]
):
    print("WE HERE YALL")
    st.session_state["last_filter"] = selected_filter
    st.session_state["last_tense_selection"] = selected_tenses
    st.session_state.pop("current_task", None)
    st.session_state.reset_input = True
    st.rerun()




# --------- TASK SETUP ---------
if "current_task" not in st.session_state:
    row, col, verb, prompt, translation = input.get_random_task(ws_solution, selected_filter, selected_cols)
    st.session_state.current_task = {
        "row": row,
        "col": col,
        "verb": verb,
        "prompt": prompt,
        "translation": translation
    }
else:
    task = st.session_state.current_task
    row = task["row"]
    col = task["col"]
    verb = task["verb"]
    prompt = task["prompt"]
    translation = task["translation"]

# If the verb has changed, trigger input reset
if verb != st.session_state.last_verb:
    st.session_state.reset_input = True
    st.session_state.last_verb = verb

# --------- UI + LOGIC ---------
if row is None:
    st.success("🎉 All verbs have been completed!")
else:
    st.subheader(f"Verb: **{verb}**")
    st.write(f"Conjugate for: **{prompt}**")

    tense = ws_solution[f"{col}1"].value
    subject = ws_solution[f"{col}2"].value
    # translation = ws_solution[f"{con.TRANSLATION_COL}"].value
    print(translation)

    # Show translation
    with st.expander("📘 Translation", expanded=False):
        st.markdown(f"**{translation}**")

    # Show plotly diagram of tense and personal pronoun    
    input.show_conjugation_position(tense, subject)

    # Show user input field
    user_input = st.text_input("Your conjugation:", key=f"user_input_{verb}")


    if st.button("Check answer"):
        st.session_state.attempts += 1
        
        # Fetch the correct answer from the solution sheet
        correct_answer = str(ws_solution[f"{col}{row}"].value).strip()

        # Save user input to the input sheet
        cell = ws_input[f"{col}{row}"] # TODO: Check correct cell referencing for UserInput sheet

        is_correct, cleaned_input = check_user_input(user_input, correct_answer, cell)

        # Compare and apply style
        if is_correct:
            st.success("✅ Correct!")
            st.session_state.attempts = 0
            st.session_state.reset_input = True
            

        else:
            st.error(f"❌ Incorrect. Try again or reveal answer.")
            if st.session_state.attempts >= 1:
                with st.expander("📖 Show correct answer"):
                    st.markdown(f"**Correct answer:** `{correct_answer}`")

            # TODO: Retrying incorrect tries empties the input cell, but here we would want to keep it
            log_incorrect_attempt(verb, tense, subject, user_input, correct_answer, log_path="error_log.csv")

        # Check if full row is complete
        filled = all(ws_input[f"{c}{row}"].value not in [None, ""] for c in con.CONJUGATION_COLS)
        if filled:
            ws_input[f"{con.STATUS_COL}{row}"].value = "True"

        wb.save(con.EXCEL_FILE)

        # # ✅ Clear the input field AFTER saving and feedback
        st.session_state.clear_input = True

if st.button("Next verb"):
    st.session_state.pop("current_task", None)
    st.session_state.reset_input = True  # ✅ sets flag
    st.rerun()


# TODO: Allow accent's to be omitted for the word to be correct? (e.g. with unidecode library)
# TODO: Set option of different modes: random verb and form / go through verb one by one in all forms / go through one tense entirely for a verb but only that one tense (and show all personal pronouns at once with input fields)
# TODO: Check if a word has been "learned" if all inputs in the UserInput Sheet are correct and then mark it as TRUE (boolean) and not "True"
# TODO: Show progress (e.g. "50/1000 verbs completed")
# TODO: Show example/practice sentences

# WRITEBACK:
# TODO: Fill (by hand!) remaining empty cells in the Excel tab "Solutions"
# TODO: Fix that the wrong answer is written back to the wrong cell in Excel somehow...
# TODO: Should previously wrong answers be overwritten? Excluded from future runs? Excluded unless you do X?


# SHIPPING:
# TODO: Include a user feedback field
# TODO: Write user signup
# TODO: Write docker file


# FURTHER IDEAS:
# TODO: Link to full table of conjugations for given verb?
# TODO: Audio playback (using an MP3 and st.audio())?
# TODO: A third tab with vocab trainer? You could for example add a button on the Practice tab that adds a word to the vocab trainer with translation